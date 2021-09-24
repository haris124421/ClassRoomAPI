'''
Created on 31-Aug-2021

@author: muhammad.haris
'''
import time
from datetime import datetime
import os
import platform
import random
import string
import names
import base64
from openpyxl import load_workbook
import openpyxl.styles.alignment as XLStyle
from openpyxl.styles import PatternFill

class CommonFunctions():
    
    
    PrereqTestCasesStatusUpdate = False # True or False
    SuccessMessage = "Test Case Passed Successfully"
    FailedMessage = "Failed"
    # global ExecutionDate,ExecutionTime,OutPutFilePath,SystemUser,WindowServer
    ExecutionDate = str(datetime.now().date())
    ExecutionTime = str(time.strftime("%H:%M:%S", time.localtime()))
    platformsystem = str(platform.system())
    platformrelease = str(platform.release())
    WindowServer = str(platformsystem+platformrelease)
    SystemUser = os.getlogin()
    
    authuser = 'ayesha.inam@amigo-software.com'
    EncrypEmail = str(base64.b64encode(authuser.encode('ascii')))[2:-1]
    authpasssword = 'Password_123'
    Domain = 'https://web.rainbow-classroom.com/provisioningapi'
    mastertoken = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJBbWlnb1NvZnR3YXJlIiwic3ViIjoiT1BDWFJBUEkiLCJlbWFpbCI6ImFzc2lzdGFuY2VAYW1pZ28tc29mdHdhcmUuY29tIiwicm9sZSI6Ik1hbnVmYWN0dXJlciIsIk1QVCI6IkZhbHNlIiwiaHR0cDovL3NjaGVtYXMubWljcm9zb2Z0LmNvbS93cy8yMDA4LzA2L2lkZW50aXR5L2NsYWltcy9pc3BlcnNpc3RlbnQiOiJUcnVlIiwiaWF0IjoxNjI3NjY0NDg0LCJodHRwOi8vc2NoZW1hcy5taWNyb3NvZnQuY29tL3dzLzIwMDgvMDYvaWRlbnRpdHkvY2xhaW1zL3ZlcnNpb24iOiJQcm9kdWN0aW9uIiwiZXhwIjoxOTczMjY0NDg0LCJhdWQiOiJPUENYUkFQSSIsIkxJSSI6IlRydWUiLCJMSURPUyI6IjYxMDQyZWI1MzIyOWJkODMyZDc3YTI3MiIsIlNJVEVDT0RFIjoiNjEwNDJlYjUzMjI5YmQ4MzJkNzdhMjcyIiwiZXhwaXJlc19hdCI6IjE5NzMyNjQ0ODQiLCJuYmYiOjE2Mjc2NjQ0ODR9.ZikeNDBgCuZ1Dnam9RyxEHWvezQVQxP8qgO9mgyonp0'
    OutPutFilePath= 'E:\\Automation\\ClassRoomAPI.xlsx'
    datainifilepath = 'E:\\Automation\\workspace\\ClassroomAPI\\Settings\\Data.ini'
    
    
  
    
    def GenrateSimpleStringLimit10(self):
        Simplestring = "".join([random.choice(string.ascii_uppercase) for x in range(6)])
        SimpleString=''+Simplestring+'test'
        return SimpleString
    
    def GenrateSimpleStringLimit50(self):
        Simplestring = "".join([random.choice(string.ascii_uppercase) for x in range(46)])
        SimpleString=''+Simplestring+'test'
        return SimpleString
    
    def GenerateEmail(self,firstName,lastName):
        
        username = firstName+lastName
        domain = "".join([random.choice(string.ascii_lowercase) for x in range(5)])
        Email = ''+username+'@'+domain+'.com'
        return Email
    
    def GenerateFirstName(self):
        first_name= names.get_first_name(gender="")
        return first_name
    def GenerateLastName(self):
        last_name= names.get_last_name()
        return last_name
    
    def GenerateCourseName(self):
        Course_name= 'CourseTest_'+self.GenrateSimpleStringLimit10()
        return Course_name
    
    def GenerateSectionName(self):
        Section_name= 'SectionTest_'+self.GenrateSimpleStringLimit10()
        return Section_name
    
    def GenerateRoleAsTeacher(self):
        role ='1'
        return role
    def GenerateRoleAsStudent(self):
        role ='2'
        
        return role
    
    def testStatus(self,Teststatus):
        Test =Teststatus
        
        return Test
    
    def GenerateRoleAsGuest(self):
        role ='3'
        
        return role
    
    def UpdateExcelTestCase(self,Sheetname,TestCaseID,URL,Parameters,data,status,starttime,resp):
        
        showcode = str(resp['responseCode'])
        print('ResponseCode: '+showcode)
        wb=load_workbook(''+CommonFunctions.OutPutFilePath+'')
        wb.sheetnames
        ws = wb[Sheetname]
        print(TestCaseID)
        first_column = ws['B']
        del Parameters["AuthToken"]
        for x in range(len(first_column)):
            if (first_column[x].value)  == TestCaseID:
                
                ws.cell(row=x+1, column=7).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                ws.cell(row=x+1 , column=7).value = 'Header \n'+str(Parameters).replace(',','\n')+'\n Body \n'+str(data).replace(',','\n')
                ws.cell(row=x+1 , column=5).value = URL
                ws.cell(row=x+1 , column=8).value = CommonFunctions.ExecutionDate
                #ws.cell(row=x+1 , column=9).value = CommonFunctions.ExecutionTime
                ws.cell(row=x+1 , column=9).value = str(time.strftime("%H:%M:%S", time.localtime()))
                ProcessingTime = float(str((time.process_time() - starttime + 2)))
                ws.cell(row=x+1 , column=11).value = ProcessingTime
                ws.cell(row=x+1 , column=13).value = CommonFunctions.SystemUser
                ws.cell(row=x+1 , column=14).value = CommonFunctions.WindowServer
                   
                if(status =='Passed'):
                    
                    ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')
                else:
                    ws.cell(row=x+1 , column=19).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type = 'solid')
                ws.cell(row=x+1 , column=19).value = status
                ws.cell(row=x+1, column=18).alignment = XLStyle.Alignment(horizontal='center', vertical='center', wrap_text=True, wrapText=True)
                ws.cell(row=x+1 , column=18).value = showcode
                ws.cell(row=x+1 , column=16).value = str(resp)
                wb.save(''+CommonFunctions.OutPutFilePath+'')
        
               
               
    